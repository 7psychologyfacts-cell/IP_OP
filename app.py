import os
import tempfile
import zipfile
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from datetime import datetime
from flask import Flask, request, render_template_string, send_file
import xlsxwriter
app = Flask(__name__)
HTML_FORM = '''
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Excel Processor | Sky Analytics</title>
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        body {
            font-family: 'Segoe UI', 'Poppins', 'Inter', system-ui, -apple-system, sans-serif;
            background: linear-gradient(145deg, #c7e9ff 0%, #9cc9e8 100%);
            min-height: 100vh;
            display: flex;
            align-items: center;
            justify-content: center;
            padding: 20px;
        }
        .card {
            max-width: 550px;
            width: 100%;
            background: rgba(255, 255, 255, 0.95);
            backdrop-filter: blur(2px);
            border-radius: 32px;
            box-shadow: 0 25px 45px -12px rgba(0, 0, 0, 0.25), 0 0 0 1px rgba(255, 255, 255, 0.5) inset;
            padding: 2rem 2rem 2.5rem;
            transition: transform 0.3s ease;
        }
        .card:hover {
            transform: translateY(-5px);
        }
        h2 {
            text-align: center;
            margin-bottom: 1.5rem;
            font-weight: 600;
            font-size: 1.8rem;
            background: linear-gradient(135deg, #1e4b6e, #2c7da0);
            background-clip: text;
            -webkit-background-clip: text;
            color: transparent;
            letter-spacing: -0.3px;
        }
        .form-group {
            margin-bottom: 1.5rem;
        }
        label {
            display: block;
            font-weight: 500;
            margin-bottom: 0.5rem;
            color: #1a4b6e;
            font-size: 0.9rem;
            letter-spacing: 0.3px;
        }
        input[type="file"],
        input[type="date"] {
            width: 100%;
            padding: 12px 16px;
            background-color: #f8fafc;
            border: 1px solid #cbd5e1;
            border-radius: 20px;
            font-size: 0.95rem;
            transition: all 0.2s ease;
            font-family: inherit;
        }
        input[type="file"]:hover,
        input[type="date"]:hover {
            border-color: #2c7da0;
        }
        input[type="file"]:focus,
        input[type="date"]:focus {
            outline: none;
            border-color: #1e4b6e;
            box-shadow: 0 0 0 3px rgba(44, 125, 160, 0.2);
            background-color: #ffffff;
        }
        input[type="file"]::file-selector-button {
            background: #e2e8f0;
            border: none;
            border-radius: 30px;
            padding: 8px 18px;
            margin-right: 16px;
            font-weight: 500;
            color: #1a4b6e;
            cursor: pointer;
            transition: 0.2s;
        }
        input[type="file"]::file-selector-button:hover {
            background: #cbd5e1;
        }
        button {
            width: 100%;
            background: linear-gradient(95deg, #1e4b6e, #2c7da0);
            border: none;
            padding: 14px;
            border-radius: 40px;
            color: white;
            font-weight: 600;
            font-size: 1rem;
            letter-spacing: 0.5px;
            cursor: pointer;
            transition: all 0.3s ease;
            margin-top: 10px;
            box-shadow: 0 4px 8px rgba(0, 0, 0, 0.05);
        }
        button:hover {
            transform: scale(1.02);
            background: linear-gradient(95deg, #123f5e, #1f6b8c);
            box-shadow: 0 10px 20px -5px rgba(0, 0, 0, 0.2);
        }
        button:active {
            transform: scale(0.98);
        }
        .footer-note {
            text-align: center;
            margin-top: 1.8rem;
            font-size: 0.75rem;
            color: #5b8cae;
            border-top: 1px solid #e2edf2;
            padding-top: 1.2rem;
        }
        @media (max-width: 480px) {
            .card {
                padding: 1.5rem;
            }
            h2 {
                font-size: 1.4rem;
            }
        }
    </style>
</head>
<body>
    <div class="card">
        <h2>☁️ Excel Processor</h2>
        <form method="post" action="/upload" enctype="multipart/form-data">
            <div class="form-group">
                <label>📁 Main File (.xlsx/.xls/.html)</label>
                <input type="file" name="main_file" accept=".xlsx,.xls,.html" required>
            </div>
            <div class="form-group">
                <label>📑 Lookup File (.xlsx)</label>
                <input type="file" name="lookup_file" accept=".xlsx" required>
            </div>
            <div class="form-group">
                <label>📅 Date (YYYY-MM-DD)</label>
                <input type="date" name="date" required>
            </div>
            <button type="submit">🚀 Process & Download</button>
        </form>
        <div class="footer-note">
            Secure • Serverless • Sky‑fast processing
        </div>
    </div>
</body>
</html>
'''
def apply_header_format(ws, header_color):
    header_fill = PatternFill("solid", fgColor=header_color)
    header_font = Font(bold=True, color="000000")
    thin = Side(style='thin', color="000000")
    all_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    max_col = ws.max_column
    max_row = ws.max_row
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = all_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).border = all_border
def apply_table_format(ws, table_color):
    header_fill = PatternFill("solid", fgColor=table_color)
    header_font = Font(bold=True, color="000000")
    row_fill_odd  = PatternFill("solid", fgColor="FFFFFF")
    row_fill_even = PatternFill("solid", fgColor="E8F7FB")
    thin = Side(style='thin', color="000000")
    all_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    max_col = ws.max_column
    max_row = ws.max_row
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = all_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in range(2, max_row + 1):
        fill = row_fill_even if row % 2 == 0 else row_fill_odd
        for col in range(1, max_col + 1):
            c = ws.cell(row=row, column=col)
            c.fill = fill
            c.border = all_border
def auto_col_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
def build_pivot_sheet(df_pivot, writer, sheet_name):
    df_pivot = df_pivot.copy()
    df_pivot.columns = df_pivot.columns.str.strip()
    def get_sort_key(ageing_str):
        if not isinstance(ageing_str, str):
            return float('inf')
        parts = ageing_str.split(') ', 1)
        if len(parts) > 1:
            val = parts[1]
            if '-' in val:
                return int(val.split('-')[0])
            elif '+' in val:
                return int(val.split('+')[0])
        return float('inf')
    ageing_order = sorted(df_pivot['Ageing'].dropna().unique(), key=get_sort_key)
    df_pivot['Ageing'] = pd.Categorical(df_pivot['Ageing'], categories=ageing_order, ordered=True)
    pivot = pd.pivot_table(
        df_pivot,
        index=['File Type', 'Unit', 'Sponsor Type'],
        columns='Ageing',
        values='Invoice No.',
        aggfunc='count',
        fill_value=0
    )
    pivot['Total'] = pivot.sum(axis=1)
    pivot = pivot[pivot['Total'] != 0]
    all_sponsors = set(df_pivot['Sponsor Type'].dropna().unique())
    has_prob = "Probably to be cancelled as per Unit" in all_sponsors
    has_srit = "SRIT Error" in all_sponsors
    if has_prob and has_srit:
        nc_cols = ['b) 3-5','c) 6-8','d) 9-15','e) 16-20','f) 21-30',
                   'g) 31-33','h) 34-60','i) 61-65','j) 66-180','k) 181-365','l) 365+']
        cy_cols = ['h) 34-60','i) 61-65','j) 66-180','k) 181-365','l) 365+']
    elif has_prob:
        nc_cols = ['c) 6-8','d) 9-15','e) 16-20','f) 21-30',
                   'g) 31-35','h) 36-60','i) 61-65','j) 66-180','k) 181-365','l) 365+']
        cy_cols = ['h) 36-60','i) 61-65','j) 66-180','k) 181-365','l) 365+']
    else:
        nc_cols = []
        cy_cols = []
    final_df = []
    for section in ['Non Cyclic', 'Mass Check OPD', 'Cyclic']:
        if section in pivot.index.get_level_values(0):
            temp = pivot.loc[[section]]
            final_df.append(temp)
            total_row = temp.sum()
            total_row = total_row.to_frame().T
            total_row.index = [(section + " Total", "", "")]
            final_df.append(total_row)
    final_df = pd.concat(final_df)
    grand_total = pivot.sum()
    grand_total = grand_total.to_frame().T
    grand_total.index = [("Total", "", "")]
    final_df = pd.concat([final_df, grand_total])
    final_df = final_df.astype(object)
    final_df[final_df == 0] = '-'
    final_df.to_excel(writer, sheet_name=sheet_name)
    workbook  = writer.book
    worksheet = writer.sheets[sheet_name]
    header_format = workbook.add_format({
        'bold': True, 'align': 'center', 'border': 1, 'bg_color': '#D9E1F2'
    })
    total_format = workbook.add_format({
        'bold': True, 'bg_color': '#D9E1F2', 'border': 1
    })
    normal_format = workbook.add_format({'border': 1})
    red_bold = workbook.add_format({
        'bold': True, 'font_color': 'red', 'border': 1
    })
    highlight_format = workbook.add_format({
        'bg_color': '#FFC7CE', 'border': 1
    })
    filetype_format = workbook.add_format({
        'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1
    })
    headers = ['File Type', 'Unit', 'Sponsor Type'] + list(final_df.columns)
    for col_num, col_name in enumerate(headers):
        worksheet.write(0, col_num, col_name, header_format)
    last_col = len(headers) - 1
    for row_num, idx in enumerate(final_df.index):
        file_type = idx[0]
        sponsor = idx[2] if len(idx) == 3 else ""
        for col in range(0, last_col + 1):
            if col < 3:
                value = idx[col] if col < len(idx) else ""
            else:
                value = final_df.iloc[row_num, col - 3]
            fmt = normal_format
            if col == 0 and file_type != "":
                fmt = filetype_format
            if "Total" in file_type:
                fmt = total_format
            if sponsor in ["Probably to be cancelled as per Unit", "SRIT Error"] and col == 2:
                fmt = red_bold
            if col >= 3 and value != '-':
                col_name = headers[col]
                if file_type == "Non Cyclic" and col_name in nc_cols:
                    fmt = highlight_format
                if file_type == "Cyclic" and col_name in cy_cols:
                    fmt = highlight_format
            worksheet.write(row_num + 1, col, value, fmt)
@app.route('/')
def index():
    return render_template_string(HTML_FORM)
@app.route('/upload', methods=['POST'])
def upload():
    # Get files and date
    main_file = request.files['main_file']
    lookup_file = request.files['lookup_file']
    input_date_str = request.form['date']
    if not main_file or not lookup_file:
        return "Missing files", 400
    # Create temporary directory for processing
    with tempfile.TemporaryDirectory() as tmpdir:
        # Save uploaded files
        main_path = os.path.join(tmpdir, main_file.filename)
        lookup_path = os.path.join(tmpdir, lookup_file.filename)
        main_file.save(main_path)
        lookup_file.save(lookup_path)
        # Read main data
        try:
            tables = pd.read_html(main_path)
            df_raw = tables[0]
        except:
            df_raw = pd.read_excel(main_path)
        # First row as header
        df_raw.columns = df_raw.iloc[0]
        df_raw = df_raw[1:]
        df_raw = df_raw.reset_index(drop=True)
        # Save temp Excel
        temp_file = os.path.join(tmpdir, "temp_output.xlsx")
        df_raw.to_excel(temp_file, index=False)
        # Date formatting in temp file (column M)
        wb_temp = openpyxl.load_workbook(temp_file)
        ws_temp = wb_temp.active
        date_col = 13  # M column
        for row in range(2, ws_temp.max_row + 1):
            cell = ws_temp.cell(row=row, column=date_col)
            if cell.value:
                try:
                    if isinstance(cell.value, str):
                        cell.value = datetime.strptime(cell.value, "%d/%m/%Y")
                except:
                    pass
                cell.number_format = 'DD-MM-YYYY'
        wb_temp.save(temp_file)
        # Read lookup file
        lookup = pd.ExcelFile(lookup_path)
        sheet1 = lookup.parse("Sheet1")
        sheet2 = lookup.parse("Sheet2")
        # Process main data
        df = pd.read_excel(temp_file)
        df = df.replace('\n', ' ', regex=True)
        if 'Unit' in df.columns:
            df = df[~df["Unit"].str.contains("Zynova", na=False)]
            df = df[~df["Unit"].str.contains("--End-", na=False)]
        df["TPA Name"] = df["TPA Name"].fillna(df["Sponsor/Insurance"])
        df = df.sort_values(by="Admission No.", na_position='last')
        # Split IP & OP
        ip_df = df[df["Admission No."].notna()].copy()
        op_df = df[df["Admission No."].isna()].copy()
        # Date
        input_date = pd.to_datetime(input_date_str)
        # Days + ageing function
        def add_days_ageing(data, op=False):
            data["Invoice Date"] = pd.to_datetime(
                data["Invoice Date"], dayfirst=True, errors='coerce'
            )
            data["Invoice Date"] = data["Invoice Date"].fillna(input_date)
            data["Days"] = (input_date - data["Invoice Date"]).dt.days
            if not op:
                conditions = [
                    (data["Days"] <= 2), (data["Days"] <= 5), (data["Days"] <= 8),
                    (data["Days"] <= 15), (data["Days"] <= 20), (data["Days"] <= 30),
                    (data["Days"] <= 33), (data["Days"] <= 60), (data["Days"] <= 65),
                    (data["Days"] <= 180), (data["Days"] <= 365)
                ]
                choices = [
                    "a) 0-2","b) 3-5","c) 6-8","d) 9-15","e) 16-20",
                    "f) 21-30","g) 31-33","h) 34-60","i) 61-65",
                    "j) 66-180","k) 181-365"
                ]
            else:
                conditions = [
                    (data["Days"] <= 2), (data["Days"] <= 5), (data["Days"] <= 8),
                    (data["Days"] <= 15), (data["Days"] <= 20), (data["Days"] <= 30),
                    (data["Days"] <= 35), (data["Days"] <= 60), (data["Days"] <= 65),
                    (data["Days"] <= 180), (data["Days"] <= 365)
                ]
                choices = [
                    "a) 0-2","b) 3-5","c) 6-8","d) 9-15","e) 16-20",
                    "f) 21-30","g) 31-35","h) 36-60","i) 61-65",
                    "j) 66-180","k) 181-365"
                ]
            data["Ageing"] = np.select(conditions, choices, default="l) 365+")
            return data
        ip_df = add_days_ageing(ip_df)
        op_df = add_days_ageing(op_df, op=True)

        # ---------------------------------------------------------------
        # File type logic (IP)
        # ---------------------------------------------------------------
        ip_df["File Type"] = ""

        ip_df.loc[
            (ip_df["Sponsor Type"] == "Clinical Research") & (ip_df["File Type"] == ""),
            "File Type"
        ] = "Cyclic"

        ip_df.loc[
            (ip_df["Sponsor Type"] == "Self Pay Credit") & (ip_df["File Type"] == ""),
            "File Type"
        ] = "Non Cyclic"

        ip_df.loc[
            (ip_df["TPA Name"].str.contains("rajas|rajasthan", case=False, na=False)) & (ip_df["File Type"] == ""),
            "File Type"
        ] = "Non Cyclic"

        ip_df.loc[
            (ip_df["Sponsor Type"] == "Government") &
            (ip_df["TPA Name"].str.contains("central gov|Central Gov", case=False, na=False)) &
            (ip_df["File Type"] == ""),
            "File Type"
        ] = "Cyclic"

        ip_df.loc[
            (ip_df["Invoice Amount"].astype(float) < 5000) & (ip_df["File Type"] == ""),
            "File Type"
        ] = "Cyclic"

        ip_df.loc[
            (ip_df["Unit"] == "Indore") & (ip_df["Sponsor Type"] != "Insurance") & (ip_df["File Type"] == ""),
            "File Type"
        ] = "Cyclic"

        ip_df.loc[
            (ip_df["Unit"] == "Surat") & (~ip_df["Sponsor Type"].isin(["Insurance","Government"])) & (ip_df["File Type"] == ""),
            "File Type"
        ] = "Cyclic"

        # Fill remaining blank File Types as Non Cyclic
        ip_df["File Type"] = ip_df["File Type"].replace("", "Non Cyclic")

        # ---------------------------------------------------------------
        # NEW RULE — Applied LAST, overrides any previously assigned File Type:
        # Unit == "SG"  AND  TPA Name is one of the 6 specified values
        # AND  Invoice Amount <= 10000  →  File Type = "Cyclic"
        # ---------------------------------------------------------------
        sg_tpa_list = [
            "PRL[CHSS]",
            "ISRO [CHSS]",
            "INDIAN OIL CORPORATION LTD.",
            "ONGC Ahmedabad",
            "IPR[CHSS]",
            "SAC[CHSS]",
        ]
        ip_df.loc[
            (ip_df["Unit"] == "SG") &
            (ip_df["TPA Name"].isin(sg_tpa_list)) &
            (ip_df["Invoice Amount"].astype(float) <= 10000),
            "File Type"
        ] = "Cyclic"
        # ---------------------------------------------------------------

        # Remarks mapping
        sheet2_unique = sheet2.drop_duplicates(subset=[sheet2.columns[1]])
        remarks_map = sheet2_unique.set_index(sheet2.columns[1])[sheet2.columns[8]].to_dict()
        ip_df["Remarks"] = ip_df["Invoice No."].map(remarks_map).fillna("-")
        op_df["Remarks"] = op_df["Invoice No."].map(remarks_map).fillna("-")
        # Update sponsor type
        def update_sponsor_type(df_in):
            def check_remark(row):
                if row["Remarks"] == "Probably to be cancelled as per Unit":
                    return "Probably to be cancelled as per Unit"
                elif row["Remarks"] != "-":
                    return "SRIT Error"
                else:
                    return row["Sponsor Type"]
            df_in["Sponsor Type"] = df_in.apply(check_remark, axis=1)
            return df_in
        ip_df = update_sponsor_type(ip_df)
        op_df = update_sponsor_type(op_df)
        # Move remarks to last column
        def move_remarks_last(df):
            df.columns = df.columns.str.strip()
            if "Remarks" in df.columns:
                cols = [c for c in df.columns if c != "Remarks"] + ["Remarks"]
                return df[cols]
            return df
        # Fill blanks
        ip_df = ip_df.fillna("-")
        op_df = op_df.fillna("-")
        num_cols = ["Policy No.", "ID No."]
        for col in num_cols:
            if col in ip_df.columns:
                ip_df[col] = pd.to_numeric(ip_df[col], errors='coerce')
            if col in op_df.columns:
                op_df[col] = pd.to_numeric(op_df[col], errors='coerce')
        # OP file type
        map_filetype = sheet1.set_index(sheet1.columns[0])[sheet1.columns[1]].to_dict()
        op_df["File Type"] = op_df["TPA Name"].map(map_filetype).fillna("NA")
        ip_df = move_remarks_last(ip_df)
        op_df = move_remarks_last(op_df)
        # Formatting helpers
        HEADER_COLOR = "FFCA70"
        TABLE_COLOR  = "51BFDA"

        # ------------------------------------------------------------------
        # Build NC / Cyclic subsets — IP
        # ------------------------------------------------------------------
        ip_nc = ip_df[(ip_df["File Type"] == "Non Cyclic") & (ip_df["Ageing"].isin([
            "b) 3-5","c) 6-8","d) 9-15","e) 16-20",
            "f) 21-30","g) 31-33","h) 34-60","i) 61-65",
            "j) 66-180","k) 181-365","l) 365+"
        ]))].sort_values("Unit")

        ip_cyc = ip_df[(ip_df["File Type"] == "Cyclic") & (ip_df["Ageing"].isin([
            "h) 34-60","i) 61-65","j) 66-180","k) 181-365","l) 365+"
        ]))].sort_values("Unit")

        # Sheet tab names include data-row count (header row excluded)
        ip_nc_sheet_name  = f"{len(ip_nc)} NC"
        ip_cyc_sheet_name = f"{len(ip_cyc)} Cyclic"

        # ------------------------------------------------------------------
        # Build NC / Cyclic subsets — OP
        # ------------------------------------------------------------------
        op_nc = op_df[(op_df["File Type"] == "Non Cyclic") & (op_df["Ageing"].isin([
            "c) 6-8","d) 9-15","e) 16-20","f) 21-30","g) 31-35",
            "h) 36-60","i) 61-65","j) 66-180","k) 181-365","l) 365+"
        ]))].sort_values("Unit")

        op_cyc = op_df[(op_df["File Type"] == "Cyclic") & (op_df["Ageing"].isin([
            "h) 36-60","i) 61-65","j) 66-180","k) 181-365","l) 365+"
        ]))].sort_values("Unit")

        op_nc_sheet_name  = f"{len(op_nc)} NC"
        op_cyc_sheet_name = f"{len(op_cyc)} Cyclic"

        # ------------------------------------------------------------------
        # Output IP
        # ------------------------------------------------------------------
        ip_out_path = os.path.join(tmpdir, "OUTPUT_IP.xlsx")
        with pd.ExcelWriter(ip_out_path, engine='xlsxwriter') as writer:
            ip_df.to_excel(writer, sheet_name="IP", index=False)
            ip_nc.to_excel(writer, sheet_name=ip_nc_sheet_name, index=False)
            ip_cyc.to_excel(writer, sheet_name=ip_cyc_sheet_name, index=False)
            build_pivot_sheet(ip_df, writer, "Pivot")

        # Apply openpyxl formatting
        wb = load_workbook(ip_out_path)
        apply_header_format(wb["IP"], HEADER_COLOR)
        auto_col_width(wb["IP"])
        apply_table_format(wb[ip_nc_sheet_name], TABLE_COLOR)
        auto_col_width(wb[ip_nc_sheet_name])
        apply_table_format(wb[ip_cyc_sheet_name], TABLE_COLOR)
        auto_col_width(wb[ip_cyc_sheet_name])
        wb.save(ip_out_path)

        # ------------------------------------------------------------------
        # Output OP
        # ------------------------------------------------------------------
        op_out_path = os.path.join(tmpdir, "OUTPUT_OP.xlsx")
        with pd.ExcelWriter(op_out_path, engine='xlsxwriter') as writer:
            op_df.to_excel(writer, sheet_name="OP", index=False)
            op_nc.to_excel(writer, sheet_name=op_nc_sheet_name, index=False)
            op_cyc.to_excel(writer, sheet_name=op_cyc_sheet_name, index=False)
            build_pivot_sheet(op_df, writer, "Pivot")

        wb2 = load_workbook(op_out_path)
        apply_header_format(wb2["OP"], HEADER_COLOR)
        auto_col_width(wb2["OP"])
        apply_table_format(wb2[op_nc_sheet_name], TABLE_COLOR)
        auto_col_width(wb2[op_nc_sheet_name])
        apply_table_format(wb2[op_cyc_sheet_name], TABLE_COLOR)
        auto_col_width(wb2[op_cyc_sheet_name])
        wb2.save(op_out_path)

        # Create ZIP of the two output files
        zip_path = os.path.join(tmpdir, "outputs.zip")
        with zipfile.ZipFile(zip_path, 'w') as zf:
            zf.write(ip_out_path, arcname="OUTPUT_IP.xlsx")
            zf.write(op_out_path, arcname="OUTPUT_OP.xlsx")
        return send_file(zip_path, as_attachment=True, download_name="processed_outputs.zip")
if __name__ == '__main__':
    app.run(debug=True)
